import csv
import sys
import MySQLdb
#import dbcontext
from datetime import datetime
import time    
ts = time.strftime('%Y-%m-%d %H:%M:%S')
created_at = datetime.now()
#db = DBContext()

from openpyxl import load_workbook
wb = load_workbook(filename = "data/Data.xlsx",use_iterators=True)
ws = wb.get_sheet_by_name(name = 'books')
db = MySQLdb.connect("localhost","prem1980","django","onlineshop" )
cursor = db.cursor()
row_no = 0
for row in ws.iter_rows():
	row_output = []
	cell_no = 0
	for cell in row:
		element = cell.value
		if type(element) is float or type(element) is int or element is None:
			if cell_no in [0,4,18]: #id,sku and category_id needs to be converted to float
				row_output.append(int(element))
			else:
				if type(element) is float:
					element = "{0:.3f}".format(element) 	# float formatting
					row_output.append(element)
				else:
					if element is None:
						element = "" # None is converted to space
					row_output.append(element)
		
		else:
			element = element.replace("'",'"')
			row_output.append(element.encode('ascii','ignore'))
		cell_no += 1
	row_output[14] = row_output[13] = ts
	print 'row_output =', row_output
	print 'len row_output =', len(row_output)
	print 'row_no =', row_no
	if row_no != 0:	
		stmt = """ Insert into Products (id,name) values ({0},'{1}')""".format(1,'test')
		stmt = """ Insert into Products values ({0},'{1}','{2}','{3}',{4},{5},{6},'{7}','{8}','{9}','{10}','{11}','{12}','{13}','{14}','{15}','{16}','{17}',{18})""".format(*row_output)
		
		print 'stmt =', stmt
		cursor.execute(stmt)
		db.commit()
	row_no += 1



i = 0
def readxls():
	with open("data/Products.xlsx",'rb') as csvfile:
		spamreader = csv.reader(csvfile,delimiter=",")
		for row in spamreader:
			if i in [0]:
			   pass
			else:
				#created_at and updated_at are set to current timestamp
				row[0] = int(row[0])
				row[4] = int(row[4])
				row[14] = row[13] = ts
				row[10] = row[10].replace('"','\\"')
				row[1] = row[1].replace('"','\\"')
				row[11] = row[11].replace('"','\\"')
				row[12] = row[12].replace('"','\\"')
				row[10]  = row[10].decode('unicode_escape').encode('ascii','ignore')
				row[1] = row[1].decode('unicode_escape').encode('ascii','ignore')
				row[11] = row[11].decode('unicode_escape').encode('ascii','ignore')
				row[12] = row[12].decode('unicode_escape').encode('ascii','ignore')
				var_string = ', '.join('?' * len(row))
				print row
				#Mysqldb does not allow parameterizzed query-- you can check using import MySQLdb and print MySQLdb.paramstyle
				#sql = '''INSERT INTO Products VALUES ({0})'''.format(var_string)
				#sql = " insert into Products values (1,'name','slug','authorname',1,10.00,20.00,'Y','Y','Y','DESCRIPTION','META','META','2008-01-01 00:00:00','2008-01-01 00:00:01','IMAGE','THUMBNAIL','IMAGE')"
				x = 0 
				output = ' '
				for each in row:
					if x in [0,4]:
						output = output + str(each) + ' , '		
					else:
						output = output + '"' + str(each) + '"' +  ' , '
					x += 1
			#		print 'output =', output
				output = output[:-2]
				print 'len of row =', len(output)
				sql = "INSERT INTO Products values (%s)" % output
				print 'sql =', sql
				#try:
				cursor.execute(sql)
				db.commit()
				#except Exception, e:
			#    print "Error %s" % (e.args[0])
			#    sys.exit(1)
		i += 1

                
              
